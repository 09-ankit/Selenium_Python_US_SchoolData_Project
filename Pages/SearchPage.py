import os
import openpyxl
from selenium.webdriver.common.by import By
from Utilities.CommonActions import CommonActions

class SearchPage:

    def __init__(self, driver):
        self.driver = driver
        self.actions = CommonActions(driver)
        self.School_Locator = "School"
        self.City_Locator = "City"
        self.Search_Button_Locator = "Search_Button"
        self.GetTotalSearchResult = "Search_Result"


    def enter_School(self, text):
        self.actions.enter_text(self.School_Locator, text)

    def click_Search_button(self):
        self.actions.click_element(self.Search_Button_Locator)

    def get_Result_Count(self):
        Result=self.actions.get_text(self.GetTotalSearchResult)
        print(Result)

    def Store_Schools_Data(self):

        Excel_File_Path = r"Resources_holder\StoreEmailID.xlsx"
        # Get the absolute path to the project directory
        project_directory = os.path.dirname(os.path.abspath(__file__))
        # Combine the project directory path with the relative path
        StudentData_File = os.path.join(project_directory, "..", Excel_File_Path)


        workbook = openpyxl.load_workbook(StudentData_File)
        sheet = workbook.active
        Click_NextButton=[16,31,45,61]
        i = 2
        row = sheet.max_row + 1

        while row <= 69:
            try:
                data = []

                # Collect data for each element in the row
                elements = self.driver.find_elements(By.XPATH,
                                                     f"//body[1]/div[1]/div[3]/table[3]/tbody[1]/tr[{i}]/td[1]/table[1]/tbody/tr/td/font")
                for element in elements:
                    data.append(element.text)

                # Divide the data into groups of 7 elements each
                data_groups = [data[j:j + 7] for j in range(0, len(data), 7)]

                # Store each group in a separate row in the Excel sheet
                for group in data_groups:
                    print(group)
                    for col, value in enumerate(group, start=1):
                        sheet.cell(row=row, column=col, value=value)
                    row += 1

            except Exception as e:
                print(f"Unable to print element texts: {e}")
            i+=2
            if row in Click_NextButton or row==Click_NextButton:
                self.driver.find_element(By.XPATH, "//a[normalize-space()='Next >>']").click()
                i = 3

        workbook.save(StudentData_File)









