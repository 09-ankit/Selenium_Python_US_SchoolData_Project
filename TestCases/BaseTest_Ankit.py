import time

import openpyxl
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Resources_Reader.ConfigReader import read_configs
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options



def clear_excel_sheet(file_path1, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_path1)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
            workbook.save(file_path1)
            workbook.close()
            print(f"Data cleared from sheet '{sheet_name}' in {file_path1}")
        else:
            print(f"Sheet '{sheet_name}' not found in {file_path1}")
    except Exception as e:
        print(f"An error occurred while clearing data: {str(e)}")





Body_Message = """Hello ,

I am Ankit Bhattad, and I come with a strong background in Automation Testing and QA
My Skills Includes - Selenium, Robot Framework , TestNG , Rest Assured, Cucumber , Postman 
Programming Language - Java and Python
My Certifications Include - ISTQB, Selenium ,Eggplant, Rest Assured
My qualifications include - B.E from GH Raisoni College Nagpur

If there are any job openings or upcoming opportunities that match my skill set, I would greatly appreciate it if you could share this information with me.
Please find attached Resume.
Contact number -8237538774


Thank you for your time.

Regards,
Ankit
"""

file_path = "F:\\AnkitNewWorkSpace\\ResuableFramwork\\src\\main\\java\\Resources\\StoreEmailID.xlsx"
workbook = openpyxl.load_workbook(file_path)
webdriver_path = Service( "F:\\selfils\\Ankit\\Newfolder\\Newdriver\\chromedriver.exe" )
option=Options()
option.add_experimental_option("debuggerAddress","localhost:9998")
driver = webdriver.Chrome(service=webdriver_path, options=option)
#driver.get("https://www.linkedin.com/login")
driver.implicitly_wait( 15 )
driver.set_page_load_timeout( 10 )
wait = WebDriverWait(driver, 10)
clear_excel_sheet(file_path, "Sheet1")




#
### *************As We Are working with Gmail And Linkdln due to security feature have to work with open Browser
###login Application
# driver.find_element(By.ID,'username').send_keys("ankitbhattad99@gmail.com")
# driver.find_element(By.ID,'password').send_keys("bhattad@148412")
# driver.find_element(By.XPATH,"(//*[text()='Sign in'])[2]").click()

#####Search 3 Time with List Values and store in excel

SearchTexts=["Automation Tester 2 years * ", "Automation Tester 2 years Nagpur *" , "QA 2 years Nagpur *","Automation Tester 2 years remote * "]
j=0
while(j<=3):
    ###Search Text And Enter
    wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@placeholder='Search']")))
    print(f"VALUE OF J IS NOW after incre{SearchTexts[j]}")
    driver.find_element(By.XPATH, "//input[@placeholder='Search']").clear()
    driver.find_element(By.XPATH,"//input[@placeholder='Search']").send_keys(SearchTexts[j])
    driver.find_element(By.XPATH,"//input[@placeholder='Search']").send_keys(Keys.RETURN)


    if(j==0):
        wait.until(EC.element_to_be_clickable((By.XPATH,"(//button[contains(@class,'search-reusables')])[1]")))

        ###Select Posts Fliter
        driver.find_element(By.XPATH,"(//button[contains(@class,'search-reusables')])[1]").click()
        time.sleep(5)

        ###Select Latest Posts
        driver.find_element(By.XPATH,"(//button[contains(@class,'search-reusables')])[2]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"//label[@for='sortBy-date_posted']")))
        driver.find_element(By.XPATH,"//label[@for='sortBy-date_posted']").click()
        driver.find_element(By.XPATH,"(//span[text()='Show results'])[1]").click()
        time.sleep(6)



                                        ###************Scroll 3 Time Page Down *************

    time.sleep(4)

    i=1
    while i<3:
        ### Scroll down vertically
        driver.execute_script("window.scrollBy(0, 11000)")
        time.sleep(10)
        ### Scroll right horizontally
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ARROW_RIGHT)
        ### Scroll left horizontally
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ARROW_LEFT)
        i+=1


    ###Capture Emails
    ListOfEmails=driver.find_elements(By.XPATH,"//a[(substring(@href, string-length(@href) - 3) = '.com' or substring(@href, string-length(@href) - 2) = '.in') and contains(@href, '@')]")
    time.sleep(7)
    print(f" Total Emails Found are --> {len(ListOfEmails)}")


    # Creating Excel File

    sheet = workbook.active

    # Find the next available row in the Excel sheet
    next_row = sheet.max_row + 1

    # Create a set to store unique email addresses
    unique_emails = set()

    # Write email to Excel
    for element in ListOfEmails:
        time.sleep(1)
        email_text = element.get_attribute("text").strip()  # Remove leading/trailing spaces
        unique_emails.add(email_text)

    j+=1

    print(f"VALUE OF J IS NOW after incre{j}")

    if(j==3):
        for unique_email in unique_emails:
            print(unique_email)
            sheet.cell(row=next_row, column=1, value=unique_email)
            next_row += 1


        # Save the workbook
        workbook.save(file_path)





### ###Get the handles of all open windows
Parent_window = driver.current_window_handle
### ###Substring to check for in the window title
substring = "Gmail"

def send_Email():
    driver.find_element(By.XPATH, "//div[text()='Compose']").click()
    driver.find_element(By.XPATH, "(//input[@placeholder='Subject'])").click()
    driver.find_element(By.XPATH, "(//input[@placeholder='Subject'])").send_keys("Resume for Automation Testing Profile")

    ### Open BCC Field
    actions = ActionChains(driver)
    actions.key_down(Keys.CONTROL)
    actions.key_down(Keys.SHIFT)
    actions.send_keys('B')
    actions.key_up(Keys.CONTROL)
    actions.key_up(Keys.SHIFT)
    actions.perform()

    ###### Select the sheet where emails are stored
    sheet = workbook["Sheet1"]

    ###### Iterate through the rows in the sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        email_to_send = row[0]  ### Assuming the email is in the first column (column A)
        driver.find_element(By.XPATH, "(//span[text()='Bcc'])/../following-sibling::td//input").send_keys(email_to_send)
        driver.find_element(By.XPATH, "(//span[text()='Bcc'])/../following-sibling::td//input").send_keys(Keys.RETURN)
        time.sleep(1)


    driver.find_element(By.XPATH, "//div[@aria-label='Message Body']").click()
    driver.find_element(By.XPATH, "//div[@aria-label='Message Body']").send_keys(Body_Message)






### ###Loop through all open windows and switch to the one with the desired title
for window_handle in driver.window_handles:
    driver.switch_to.window(window_handle)
    if substring in driver.title:
        time.sleep(2)
        send_Email()






