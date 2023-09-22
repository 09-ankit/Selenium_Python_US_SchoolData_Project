from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string

class Xls_Reader:
    def __init__(self, path):
        self.path = path
        try:
            self.workbook =load_workbook(path)
            self.sheet = self.workbook.active
        except Exception as e:
            print("Error: ", e)

    def getRowCount(self, sheetName):
        sheet = self.workbook[sheetName]
        return sheet.max_row

    def read_cell_data( self,sheet_name, column_name, row_number):
        # Load the Excel workbook
        workbook = load_workbook( self.path )

        try:
            # Choose a specific sheet by name
            sheet = workbook[sheet_name]

            # Get the column index based on column name
            column_index = column_index_from_string( column_name )

            # Access cell value based on row number and column index
            cell_value = sheet.cell( row=row_number, column=column_index ).value
            return cell_value

        finally:
            # Close the workbook
            workbook.close()

    def setCellData(self, sheetName, colName, rowNum, data, url=None):
        try:
            sheet = self.workbook[sheetName]
            if rowNum <= 0:
                return False
            col_Num = -1
            for i in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=i)
                if cell.value.strip() == colName:
                    col_Num = i
                    break
            if col_Num == -1:
                return False
            cell = sheet.cell(row=rowNum, column=col_Num)
            cell.value = data

            if url is not None:
                image = Image(url)
                sheet.add_image(image, f"{get_column_letter(col_Num)}{rowNum}")

            self.workbook.save(self.path)
            return True
        except Exception as e:
            print("Error: ", e)
            return False

    def addSheet(self, sheetname):
        try:
            self.workbook.create_sheet(sheetname)
            self.workbook.save(self.path)
            return True
        except Exception as e:
            print("Error: ", e)
            return False

    def removeSheet(self, sheetName):
        try:
            sheet = self.workbook[sheetName]
            self.workbook.remove(sheet)
            self.workbook.save(self.path)
            return True
        except Exception as e:
            print("Error: ", e)
            return False

    def add_column(self, sheet_name, col_name):
        sheet = self.workbook[sheet_name]
        sheet.insert_cols( sheet.max_column + 1 )
        cell = sheet.cell( row=1, column=sheet.max_column )
        cell.value = col_name
        fill = PatternFill( start_color="FFC7CE", end_color="FFC7CE", fill_type="solid" )
        cell.fill = fill
        self.workbook.save( self.path )
        return True

    def remove_column(self, sheet_name, col_num):
        sheet = self.workbook[sheet_name]
        sheet.delete_cols( col_num )
        self.workbook.save( self.path )
        return True

    def is_sheet_exist(self, sheet_name):
        return sheet_name in self.workbook.sheetnames

    def get_column_count(self, sheet_name):
        sheet = self.workbook[sheet_name]
        return sheet.max_column
